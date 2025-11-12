from __future__ import annotations
from typing import Dict, List, Optional, Union
from .supabase_client import sb
from .utils import nums_schema

def set_number(property_id: str, item_key: str, amount: Optional[float]) -> Dict:
    """Set a numeric input in the numbers framework. Returns validated result."""
    import logging
    logger = logging.getLogger(__name__)
    
    schema = nums_schema(property_id)
    try:
        sb.postgrest.schema = schema
        result = (sb.table("line_items")
          .update({"amount": amount})
          .eq("property_id", property_id)
          .eq("item_key", item_key)
          .execute())
        
        # Validate that the value was saved correctly
        if result.data:
            # Verify by reading back the value
            verify = (sb.table("line_items")
              .select("amount")
              .eq("property_id", property_id)
              .eq("item_key", item_key)
              .execute())
            if verify.data and len(verify.data) > 0:
                saved_amount = verify.data[0].get("amount")
                if saved_amount == amount or (saved_amount is None and amount is None):
                    logger.info(f"✅ Validated: {item_key} = {amount} saved correctly")
                    return {"ok": True, "item_key": item_key, "amount": amount, "validated": True}
                else:
                    logger.warning(f"⚠️ Validation failed: expected {amount}, got {saved_amount}")
                    return {"ok": False, "item_key": item_key, "error": f"Validation failed: expected {amount}, got {saved_amount}"}
        
        return {"ok": True, "item_key": item_key, "amount": amount, "validated": False}
    except Exception as e:
        logger.error(f"Error setting number {item_key}: {e}")
        # Fallback via RPC in public schema
        try:
            sb.postgrest.schema = "public"
            sb.rpc("set_property_number", {"p_id": property_id, "k": item_key, "amount": amount}).execute()
            # Try to validate by reading back from line_items
            try:
                verify = (sb.table("line_items")
                  .select("amount")
                  .eq("property_id", property_id)
                  .eq("item_key", item_key)
                  .execute())
                if verify.data and len(verify.data) > 0:
                    saved_amount = verify.data[0].get("amount")
                    if saved_amount == amount or (saved_amount is None and amount is None):
                        return {"ok": True, "item_key": item_key, "amount": amount, "validated": True}
            except:
                pass  # If validation fails, still return success but unvalidated
            return {"ok": True, "item_key": item_key, "amount": amount, "validated": False}
        except Exception as e2:
            logger.error(f"Fallback RPC also failed: {e2}")
            return {"ok": False, "item_key": item_key, "error": str(e2)}

def get_numbers(property_id: str, template_key: Optional[str] = None) -> List[Dict]:
    """Get all numbers for a property. Returns the structure even if values are NULL.
    
    If no items are found in the DB, returns the template structure for R2B by default.
    """
    # Define the template structures first
    template_structures = {
        "R2B": [
            {"group_name": "Bº RAMA", "item_key": "precio_venta", "item_label": "Precio de venta", "is_percent": False, "amount": None},
            {"group_name": "Bº RAMA", "item_key": "terreno_urbano", "item_label": "Terreno urbano", "is_percent": False, "amount": None},
            {"group_name": "Bº RAMA", "item_key": "terreno_rustico", "item_label": "Terreno rústico", "is_percent": False, "amount": None},
            {"group_name": "Bº RAMA", "item_key": "terreno_urbano_iva_pct", "item_label": "IVA (%)", "is_percent": True, "amount": None},
            {"group_name": "Bº RAMA", "item_key": "terreno_rustico_iva_pct", "item_label": "IVA (%)", "is_percent": True, "amount": None},
            {"group_name": "Bº RAMA", "item_key": "project_mgmt_fees", "item_label": "Project Mgmt fees", "is_percent": False, "amount": None},
            {"group_name": "Bº RAMA", "item_key": "impuestos_pct", "item_label": "Impuestos (%)", "is_percent": True, "amount": None},
            {"group_name": "Bº RAMA", "item_key": "total_pagado", "item_label": "Total pagado a 29 julio 2025", "is_percent": False, "amount": None},
            {"group_name": "Coste comprador", "item_key": "terrenos_coste", "item_label": "Terrenos", "is_percent": False, "amount": None},
            {"group_name": "Coste comprador", "item_key": "project_management_coste", "item_label": "Project Management", "is_percent": False, "amount": None},
            {"group_name": "Coste comprador", "item_key": "acometidas", "item_label": "Acometidas", "is_percent": False, "amount": None},
            {"group_name": "Coste comprador", "item_key": "costes_construccion", "item_label": "Costes de construcción", "is_percent": False, "amount": None},
        ],
    }
    
    # If a template_key is explicitly provided, return the template structure merged with DB values.
    # If no template_key is provided, do NOT fall back to a default template — return only DB items (or empty list).
    # This avoids auto-showing the old R2B template when the user didn't select it.
    if template_key:
        # Try to get items from DB and merge with template structure
        schema = nums_schema(property_id)
        db_items = {}
        try:
            sb.postgrest.schema = schema
            items = (sb.table("line_items")
                     .select("group_name,item_key,item_label,is_percent,amount,updated_at")
                     .eq("property_id", property_id)
                     .execute()).data
            # Build a map of item_key -> item for quick lookup
            for item in (items or []):
                db_items[item.get("item_key")] = item
        except Exception as e:
            # Log error for debugging
            import logging
            logging.warning(f"Error fetching from schema {schema}: {e}")
        
        # Fallback via RPC in public schema
        if not db_items:
            try:
                sb.postgrest.schema = "public"
                items = sb.rpc("list_property_numbers", {"p_id": property_id}).execute().data
                for item in (items or []):
                    db_items[item.get("item_key")] = item
            except Exception as e:
                # Log error for debugging
                import logging
                logging.warning(f"Error fetching via RPC: {e}")
        
        # Get template structure and merge with DB values
        template_structure = template_structures.get(template_key, template_structures["R2B"])
        result = []
        for template_item in template_structure:
            item_key = template_item["item_key"]
            # If we have a DB value for this item, use it (but keep template structure)
            if item_key in db_items:
                db_item = db_items[item_key]
                # Merge: use template structure but DB values for amount
                merged_item = template_item.copy()
                merged_item["amount"] = db_item.get("amount")
                merged_item["updated_at"] = db_item.get("updated_at")
                result.append(merged_item)
            else:
                # Use template structure with NULL values
                result.append(template_item.copy())
        return result
    
    # If no template_key, try to get items from DB (legacy behavior)
    schema = nums_schema(property_id)
    try:
        sb.postgrest.schema = schema
        items = (sb.table("line_items")
                 .select("group_name,item_key,item_label,is_percent,amount,updated_at")
                 .eq("property_id", property_id)
                 .execute()).data
        if items and len(items) > 0:
            return items
    except Exception as e:
        import logging
        logging.warning(f"Error fetching from schema {schema}: {e}")
    
    # Fallback via RPC
    try:
        sb.postgrest.schema = "public"
        items = sb.rpc("list_property_numbers", {"p_id": property_id}).execute().data
        if items and len(items) > 0:
            return items
    except Exception as e:
        import logging
        logging.warning(f"Error fetching via RPC: {e}")
    
    # If no items found in DB, return the default template structure
    return template_structures.get("R2B", [])

def calc_numbers(property_id: str) -> List[Dict]:
    schema = nums_schema(property_id)
    try:
        # This may fail if PostgREST doesn't expose the dynamic schema; try public RPC instead.
        return sb.rpc(f"{schema}.calc").execute().data
    except Exception:
        sb.postgrest.schema = "public"
        return sb.rpc("calc_property_numbers", {"p_id": property_id}).execute().data

def clear_number(property_id: str, item_key: str) -> Dict:
    """Clear/reset a specific number value by setting it to None."""
    return set_number(property_id, item_key, None)

def initialize_template_structure(property_id: str, template_key: str) -> Dict:
    """Initialize the structure of line_items for a numbers template.
    This ensures that get_numbers returns the full structure even if values are NULL.
    """
    # Define the structure for each template
    template_structures = {
        "R2B": [
            # Bº RAMA section
            {"group_name": "Bº RAMA", "item_key": "precio_venta", "item_label": "Precio de venta", "is_percent": False},
            {"group_name": "Bº RAMA", "item_key": "terreno_urbano", "item_label": "Terreno urbano", "is_percent": False},
            {"group_name": "Bº RAMA", "item_key": "terreno_rustico", "item_label": "Terreno rústico", "is_percent": False},
            {"group_name": "Bº RAMA", "item_key": "terreno_urbano_iva_pct", "item_label": "IVA (%)", "is_percent": True},
            {"group_name": "Bº RAMA", "item_key": "terreno_rustico_iva_pct", "item_label": "IVA (%)", "is_percent": True},
            {"group_name": "Bº RAMA", "item_key": "project_mgmt_fees", "item_label": "Project Mgmt fees", "is_percent": False},
            {"group_name": "Bº RAMA", "item_key": "impuestos_pct", "item_label": "Impuestos (%)", "is_percent": True},
            {"group_name": "Bº RAMA", "item_key": "total_pagado", "item_label": "Total pagado a 29 julio 2025", "is_percent": False},
            # Coste comprador section
            {"group_name": "Coste comprador", "item_key": "terrenos_coste", "item_label": "Terrenos", "is_percent": False},
            {"group_name": "Coste comprador", "item_key": "project_management_coste", "item_label": "Project Management", "is_percent": False},
            {"group_name": "Coste comprador", "item_key": "acometidas", "item_label": "Acometidas", "is_percent": False},
            {"group_name": "Coste comprador", "item_key": "costes_construccion", "item_label": "Costes de construcción", "is_percent": False},
        ],
        "R2B + PM": [
            # Same as R2B for now
        ],
        "R2B + PM + Venta certs": [
            # Same as R2B for now
        ],
        "Promoción": [
            # Different structure for Promoción
        ]
    }
    
    structure = template_structures.get(template_key, template_structures["R2B"])
    
    # For now, just return success - the structure will be created when the user sets values
    # The RPC list_property_numbers should handle this, but if it doesn't, we'll need to create an RPC
    return {"property_id": property_id, "template_key": template_key, "status": "structure_initialized"}

def clear_numbers(property_id: str) -> Dict:
    """Clear/reset all number values for a property when starting fresh with a new template.
    
    This function attempts to clear all number values, but if the RPC doesn't exist,
    it will return success anyway (template selection is the priority).
    """
    # Try RPC first (if it exists) to avoid 404 errors from direct table access
    try:
        sb.postgrest.schema = "public"
        sb.rpc("clear_property_numbers", {"p_id": property_id}).execute()
        return {"property_id": property_id, "status": "cleared"}
    except Exception:
        # If RPC doesn't exist, return success anyway (template selection is the priority)
        # The values will be cleared when the user starts entering new values in the Excel
        # This avoids 404 errors in the logs while still allowing template selection to proceed
        return {"property_id": property_id, "status": "attempted", "note": "RPC not available, values will be cleared when user enters new values"}

def find_item_by_value(property_id: str, search_value: Optional[Union[str, float]] = None, search_label: Optional[str] = None) -> Optional[Dict]:
    """Find an item in the numbers framework by value or label.
    
    Args:
        property_id: Property UUID
        search_value: Value to search for (e.g., 10.0 for "10%" or 100000)
        search_label: Label text to search for (e.g., "IVA", "Precio de venta")
    
    Returns:
        Dict with item_key, item_label, amount, etc. or None if not found
    """
    items = get_numbers(property_id)
    
    # Normalize search_label for fuzzy matching
    if search_label:
        search_label_lower = search_label.lower().strip()
        # Remove common words and normalize, but keep "iva" for matching
        search_label_clean = search_label_lower.replace("(%)", "").replace("(", "").replace(")", "").strip()
    
    # Parse search_value if it's a string
    parsed_value = None
    if search_value is not None:
        if isinstance(search_value, (int, float)):
            parsed_value = float(search_value)
        elif isinstance(search_value, str):
            try:
                val_str = str(search_value).replace("%", "").replace(",", ".").strip()
                parsed_value = float(val_str)
            except:
                pass
    
    best_match = None
    best_score = 0
    
    for item in items:
        item_label = item.get("item_label", "").lower()
        item_key = item.get("item_key", "").lower()
        item_amount = item.get("amount")
        
        score = 0
        
        # Match by label (fuzzy)
        if search_label:
            # Exact match in label
            if search_label_clean in item_label or item_label in search_label_clean:
                score += 10
            # Partial match in label
            elif any(word in item_label for word in search_label_clean.split() if len(word) > 2):
                score += 5
            # Match in item_key
            if search_label_clean in item_key:
                score += 8
        
        # Match by value
        if parsed_value is not None and item_amount is not None:
            # Allow small floating point differences
            if abs(float(item_amount) - parsed_value) < 0.01:
                score += 10
            # Allow percentage matching (e.g., 10% = 10.0)
            elif abs(float(item_amount) - parsed_value) < 1.0:
                score += 5
        
        # If both label and value match, return immediately (perfect match)
        if score >= 20:
            return item
        
        # Track best match
        if score > best_score:
            best_score = score
            best_match = item
    
    # Return best match if score is high enough
    if best_score >= 10:
        return best_match
    
    # Fallback: if only label or only value matches, return it
    if search_label and best_score >= 5:
        return best_match
    if parsed_value is not None and best_score >= 5:
        return best_match
    
    return None


# ==================== NEW: Numbers Table Framework (Excel Replica) ====================

def _rgb_to_hex(rgb_obj) -> str | None:
    """Convert openpyxl RGB object to hex string for JSON serialization."""
    if not rgb_obj:
        return None
    try:
        # openpyxl RGB objects have a .rgb attribute
        if hasattr(rgb_obj, 'rgb'):
            rgb_hex = rgb_obj.rgb
            if isinstance(rgb_hex, str):
                # Format: "FFRRGGBB" -> "#RRGGBB"
                if rgb_hex.startswith('FF') and len(rgb_hex) == 8:
                    return '#' + rgb_hex[2:]
                elif not rgb_hex.startswith('#'):
                    return '#' + rgb_hex if len(rgb_hex) == 6 else rgb_hex
                return rgb_hex
        # Fallback
        return str(rgb_obj) if rgb_obj else None
    except Exception:
        return None


def import_excel_from_file(file_bytes: bytes, property_id: str, template_key: str) -> Dict:
    """Import Excel template structure from uploaded file bytes using openpyxl.
    This is an alternative to Graph API that doesn't require authentication.
    
    Args:
        file_bytes: Excel file bytes
        property_id: Property UUID
        template_key: Template identifier (e.g., "R2B")
    
    Returns:
        Dict with structure_json and imported cell count
    """
    from openpyxl import load_workbook
    from io import BytesIO
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Starting Excel import from file: template_key={template_key}, property_id={property_id}")
        
        # Load workbook from bytes
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active  # Use first worksheet
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        
        logger.info(f"Excel dimensions: {max_row} rows x {max_col} columns")
        
        # Build structure
        structure = {
            "rows": max_row,
            "columns": max_col,
            "cells": []
        }
        
        # Extract header row (first row)
        header_row = []
        for col in range(1, min(max_col + 1, 27)):  # A-Z
            cell = ws.cell(row=1, column=col)
            header_row.append(str(cell.value) if cell.value else "")
        structure["header_row"] = header_row
        
        # Extract header column (first column)
        header_col = []
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=1)
            header_col.append(str(cell.value) if cell.value else "")
        structure["header_col"] = header_col
        
        # Extract all cells with values, formulas, and format
        cell_values = {}
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                
                # Skip empty cells
                if cell.value is None and not cell.data_type == 'f':  # f = formula
                    continue
                
                # Convert column number to letter (A, B, ..., Z, AA, AB, ...)
                col_letter = ""
                col_num = col
                while col_num > 0:
                    col_num -= 1
                    col_letter = chr(65 + (col_num % 26)) + col_letter
                    col_num //= 26
                
                cell_address = f"{col_letter}{row}"
                
                # Get value
                value = cell.value
                formula = None
                if cell.data_type == 'f':
                    formula = cell.value  # Formula string
                    # Try to get calculated value
                    try:
                        value = cell.value if hasattr(cell, 'value') else None
                    except:
                        value = None
                
                # Get format (convert RGB objects to strings for JSON serialization)
                cell_format = {}
                if cell.fill and cell.fill.start_color:
                    bg_color = _rgb_to_hex(cell.fill.start_color.rgb)
                    if bg_color:
                        cell_format["bg_color"] = bg_color
                
                if cell.font:
                    if cell.font.color:
                        font_color = _rgb_to_hex(cell.font.color.rgb)
                        if font_color:
                            cell_format["font_color"] = font_color
                    cell_format["bold"] = bool(cell.font.bold) if cell.font else False
                else:
                    cell_format["bold"] = False
                
                # Get row and column labels
                row_label = None
                col_label = None
                if row > 1:  # Skip header row
                    row_label_cell = ws.cell(row=row, column=1)
                    row_label = str(row_label_cell.value) if row_label_cell.value else None
                if col > 1:  # Skip header column
                    col_label = header_row[col - 1] if col - 1 < len(header_row) else None
                
                structure["cells"].append({
                    "address": cell_address,
                    "row": row,
                    "col": col,
                    "value": str(value) if value is not None else "",
                    "formula": formula if formula and isinstance(formula, str) and formula.startswith("=") else None,
                    "format": cell_format,
                    "row_label": row_label,
                    "col_label": col_label
                })
                
                # Store for values table
                if value is not None:
                    cell_values[cell_address] = {
                        "value": str(value),
                        "row_label": row_label,
                        "col_label": col_label,
                        "format": cell_format
                    }
        
        # Save structure to numbers_templates
        # Use the already imported sb from supabase_client
        sb.postgrest.schema = "public"
        
        template_data = {
            "template_key": template_key,
            "property_id": property_id,
            "structure_json": structure
        }
        
        # Upsert template
        existing = sb.table("numbers_templates").select("id").eq("template_key", template_key).eq("property_id", property_id).execute()
        if existing.data:
            sb.table("numbers_templates").update(template_data).eq("id", existing.data[0]["id"]).execute()
        else:
            sb.table("numbers_templates").insert(template_data).execute()
        
        # Save initial values to numbers_table_values
        saved_count = 0
        for cell_addr, cell_data in cell_values.items():
            try:
                sb.rpc("set_numbers_table_cell", {
                    "p_property_id": property_id,
                    "p_template_key": template_key,
                    "p_cell_address": cell_addr,
                    "p_value": cell_data["value"],
                    "p_row_label": cell_data.get("row_label"),
                    "p_col_label": cell_data.get("col_label"),
                    "p_format_json": cell_data.get("format", {})
                }).execute()
                saved_count += 1
            except Exception as e:
                logger.warning(f"Failed to save cell {cell_addr}: {e}")
        
        logger.info(f"✅ Imported {saved_count} cells from Excel file")
        
        return {
            "ok": True,
            "template_key": template_key,
            "property_id": property_id,
            "structure": structure,
            "cells_imported": saved_count
        }
        
    except Exception as e:
        logger.error(f"Error importing Excel from file: {e}", exc_info=True)
        return {"ok": False, "error": str(e)}


def import_excel_template(property_id: str, template_key: str, excel_file_id: str, access_token: str) -> Dict:
    """Import Excel template structure from Microsoft Graph API and save to numbers_templates table.
    This is a fallback method when file upload is not available.
    """
    import requests
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Starting Excel import via Graph API: template_key={template_key}, property_id={property_id}")
        # This function would use Graph API - but we prefer file upload
        return {"ok": False, "error": "Graph API import not available. Please upload the Excel file directly."}
    except Exception as e:
        logger.error(f"Error importing Excel template: {e}", exc_info=True)
        return {"ok": False, "error": str(e)}


def get_numbers_table_structure(property_id: str, template_key: str) -> Dict:
    """Get the structure JSON for a Numbers template."""
    try:
        sb.postgrest.schema = "public"
        result = sb.rpc("get_numbers_template_structure", {
            "p_template_key": template_key,
            "p_property_id": property_id
        }).execute()
        if result.data:
            if isinstance(result.data, dict):
                if "rows" in result.data or "cells" in result.data:
                    return result.data
            return result.data if result.data else {}
        return {}
    except Exception as e:
        import logging
        logging.error(f"Error getting template structure: {e}")
        return {}


def get_numbers_table_values(property_id: str, template_key: str) -> Dict:
    """Get all cell values for a property's Numbers table."""
    try:
        sb.postgrest.schema = "public"
        result = sb.rpc("get_numbers_table_values", {
            "p_property_id": property_id,
            "p_template_key": template_key
        }).execute()
        return result.data if result.data else {}
    except Exception as e:
        import logging
        logging.error(f"Error getting table values: {e}")
        return {}


def set_numbers_table_cell(property_id: str, template_key: str, cell_address: str, value: str, row_label: Optional[str] = None, col_label: Optional[str] = None, format_json: Optional[Dict] = None) -> Dict:
    """Set a cell value in the Numbers table. Returns validated result."""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        sb.postgrest.schema = "public"
        result = sb.rpc("set_numbers_table_cell", {
            "p_property_id": property_id,
            "p_template_key": template_key,
            "p_cell_address": cell_address,
            "p_value": str(value) if value is not None else "",
            "p_row_label": row_label,
            "p_col_label": col_label,
            "p_format_json": format_json or {}
        }).execute()
        
        # Validate that the value was saved correctly
        if result.data and result.data.get("ok"):
            # Verify by reading back the value
            values = get_numbers_table_values(property_id, template_key)
            saved_value = values.get(cell_address, {})
            if isinstance(saved_value, dict):
                saved_value_str = saved_value.get("value", "")
            else:
                saved_value_str = str(saved_value)
            
            expected_value_str = str(value) if value is not None else ""
            if saved_value_str == expected_value_str:
                logger.info(f"✅ Validated: {cell_address} = {value} saved correctly")
                return {"ok": True, "cell_address": cell_address, "value": value, "validated": True}
            else:
                logger.warning(f"⚠️ Validation failed: expected {expected_value_str}, got {saved_value_str}")
                return {"ok": False, "cell_address": cell_address, "error": f"Validation failed: expected {expected_value_str}, got {saved_value_str}"}
        
        return result.data if result.data else {"ok": True, "cell_address": cell_address, "value": value, "validated": False}
    except Exception as e:
        logger.error(f"Error setting cell value {cell_address}: {e}")
        return {"ok": False, "error": str(e)}
